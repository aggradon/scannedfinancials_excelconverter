import os
from openai import OpenAI
import pandas as pd
from PIL import Image
import io
import base64
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Replace 'your-api-key-here' with your actual OpenAI API key
client = OpenAI(api_key='API Key Here')

def process_image(image_path):
    with Image.open(image_path) as img:
        byte_arr = io.BytesIO()
        img.save(byte_arr, format='PNG')
        byte_arr = byte_arr.getvalue()

    prompt = """
    Context:
    Your primary role is to assist users by interpreting images of financial statements, specifically the three main Balance sheets, Income st, and Cash Flow st. You are designed to process and convert financial information into a user-friendly, tabulated format (copyable into excel). Extract ALL relevant data from these documents and organize this information into a standardized, tab-delimited table format. Ensure accuracy and clarity in data, adapting line items to maintain consistency with the standardized format. DO NOT make assumptions about the data. 

    When presented with the images of these financial st, carefully analyze them to identify and extract financial data accurately. If the provided image is unclear return UNCLEAR IMAGE
    ------
    Specific Instructions for all financial statements:
    1. where a "-" is used to represent a 0, put 0.
    2. remove the "notes" column.
    3. the header row titled "Financial Year" should contain FY19, FY20, and so on to denote the fiscal year arranged chronologically from LEFT TO RIGHT.
    4. combine the data sets when provided with multiple years worth of data.
    5. For Income Statements ensure the line items follow these line names where possible: 
    a. Revenue, 
    b. COGS, 
    c. GP, 
    d. Selling Expenses, 
    e. General & Admin Expenses, 
    f. Other Income, 
    g. Other Expense, 
    h. Finance Income, 
    i. Finance Expense, 
    j. Profit Before Tax, 
    k. Tax, 
    l. Profit After Tax. 
    Ensure ALL expense rows that do not fit the line item buckets listed above are ALSO added in, and remove any operating profit lines. 

    Return only the table
    """

    response = client.chat.completions.create(
        model="gpt-4-vision-preview",
        messages=[
            {
                "role": "user",
                "content": [
                    {"type": "text", "text": prompt},
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/png;base64,{base64.b64encode(byte_arr).decode('utf-8')}",
                        },
                    },
                ],
            }
        ],
        max_tokens=1000,
    )
    
    return response.choices[0].message.content

def update_pnl_data(pnl_data, transcription):
    lines = transcription.strip().split('\n')
    headers = lines[0].split('\t')
    years = headers[1:]  # Assuming the first column is for line items

    for line in lines[1:]:
        items = line.split('\t')
        line_item = items[0]
        values = items[1:]

        if line_item not in pnl_data:
            pnl_data[line_item] = {}

        for year, value in zip(years, values):
            pnl_data[line_item][year] = value

    return pnl_data

def export_to_excel(pnl_data, output_file='pnl_output.xlsx'):
    # Convert pnl_data to a pandas DataFrame
    df = pd.DataFrame(pnl_data).T.reset_index()
    df.columns.name = None
    df = df.rename(columns={'index': 'Line Item'})

    # Sort columns to ensure 'Line Item' is first, followed by years in order
    year_columns = [col for col in df.columns if col != 'Line Item']
    year_columns.sort()
    df = df[['Line Item'] + year_columns]

    # Export to Excel
    writer = pd.ExcelWriter(output_file, engine='openpyxl')
    df.to_excel(writer, sheet_name='PnL', index=False)

    # Get the workbook and the worksheet
    workbook = writer.book
    worksheet = writer.sheets['PnL']

    # Define styles
    header_font = Font(bold=True)
    centered_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Apply styles to header row
    for cell in worksheet[1]:
        cell.font = header_font
        cell.alignment = centered_alignment
        cell.border = border

    # Apply styles to data cells and adjust column widths
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='center')
            cell.border = border
            if cell.column > 1:  # For numeric columns
                cell.number_format = '#,##0'

    # Adjust column widths
    for column in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    # Freeze the top row and first column
    worksheet.freeze_panes = 'B2'

    writer.close()
    print(f"Excel file '{output_file}' has been created.")

def main():
    image_files = [f for f in os.listdir('input_images') if f.endswith(('.png', '.jpg', '.jpeg'))]
    image_files.sort()  # Ensure files are processed in order

    pnl_data = {}
    all_years = set()

    for image_file in image_files:
        transcription = process_image(os.path.join('input_images', image_file))
        if transcription == "UNCLEAR IMAGE":
            print(f"Warning: Unclear image detected for {image_file}")
            continue
        
        # Update pnl_data with the new transcription
        new_pnl_data = update_pnl_data({}, transcription)
        
        # Merge new_pnl_data into pnl_data
        for line_item, year_data in new_pnl_data.items():
            if line_item not in pnl_data:
                pnl_data[line_item] = {}
            pnl_data[line_item].update(year_data)
            all_years.update(year_data.keys())
        
        # After each image, send the current state back to GPT for verification and completion
        current_state = format_current_state(pnl_data, all_years)
        verified_state = verify_and_complete_state(current_state, image_file)
        pnl_data = parse_verified_state(verified_state)

    # Export to Excel
    export_to_excel(pnl_data)

def format_current_state(pnl_data, all_years):
    years_list = sorted(list(all_years))
    lines = ['Line Item\t' + '\t'.join(years_list)]
    for line_item, year_data in pnl_data.items():
        values = [year_data.get(year, '') for year in years_list]
        lines.append(f"{line_item}\t" + '\t'.join(values))
    return '\n'.join(lines)

def verify_and_complete_state(current_state, image_file):
    prompt = f"""
    Here's the current state of the PnL data after processing {image_file}:

    {current_state}

    Please verify this data and complete any missing information based on your analysis of all images processed so far. 
    Ensure all line items are present and consistent across years. 
    If a value is truly missing or not applicable for a specific year, use '0' or 'N/A' as appropriate.
    Return the complete, verified table in the same tab-delimited format.
    """

    response = client.chat.completions.create(
        model="gpt-4-turbo-preview",
        messages=[
            {"role": "system", "content": "You are a financial data analyst assistant."},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1500
    )
    
    return response.choices[0].message.content

def parse_verified_state(verified_state):
    lines = verified_state.strip().split('\n')
    headers = lines[0].split('\t')
    years = headers[1:]
    pnl_data = {}

    for line in lines[1:]:
        items = line.split('\t')
        line_item = items[0]
        values = items[1:]

        pnl_data[line_item] = {year: value for year, value in zip(years, values)}

    return pnl_data

if __name__ == "__main__":
    main()